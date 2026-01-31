"""
Excel解析服务
职责：解析用户上传的Excel模板文件
注意：处理编码问题和去除前后空格
"""
import io
import chardet
from typing import Dict, Any, List, Optional, BinaryIO
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


class ExcelParserService:
    """Excel解析服务"""

    # 支持的参数列名映射（支持中英文）
    PARAM_COLUMN_MAPPING = {
        # 英文
        'key': 'paramKey',
        'paramkey': 'paramKey',
        'param_key': 'paramKey',
        'name': 'paramName',
        'paramname': 'paramName',
        'param_name': 'paramName',
        'value': 'value',
        'unit': 'unit',
        'remark': 'remark',
        # 中文
        '参数键': 'paramKey',
        '参数key': 'paramKey',
        '键名': 'paramKey',
        '参数名': 'paramName',
        '参数名称': 'paramName',
        '名称': 'paramName',
        '值': 'value',
        '参数值': 'value',
        '单位': 'unit',
        '备注': 'remark',
        '说明': 'remark',
    }

    # 支持的输出列名映射
    OUTPUT_COLUMN_MAPPING = {
        'key': 'outputKey',
        'outputkey': 'outputKey',
        'output_key': 'outputKey',
        'code': 'outputCode',
        'outputcode': 'outputCode',
        'output_code': 'outputCode',
        'name': 'outputName',
        'outputname': 'outputName',
        'output_name': 'outputName',
        'unit': 'unit',
        # 中文
        '输出键': 'outputKey',
        '输出key': 'outputKey',
        '输出代码': 'outputCode',
        '输出名': 'outputName',
        '输出名称': 'outputName',
        '单位': 'unit',
    }

    def parse_param_excel(
        self,
        file_content: BinaryIO,
        sheet_name: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        解析参数Excel文件

        Args:
            file_content: 文件内容（二进制流）
            sheet_name: 工作表名称，默认第一个

        Returns:
            {
                'success': bool,
                'params': [...],
                'errors': [...],
                'warnings': [...]
            }
        """
        try:
            wb = load_workbook(filename=file_content, read_only=True, data_only=True)
            ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

            return self._parse_sheet(ws, self.PARAM_COLUMN_MAPPING, 'paramKey')

        except InvalidFileException:
            return {
                'success': False,
                'params': [],
                'errors': [{'type': 'invalid_file', 'message': '无效的Excel文件格式'}],
                'warnings': []
            }
        except Exception as e:
            return {
                'success': False,
                'params': [],
                'errors': [{'type': 'parse_error', 'message': f'解析失败: {str(e)}'}],
                'warnings': []
            }

    def parse_output_excel(
        self,
        file_content: BinaryIO,
        sheet_name: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        解析输出Excel文件

        Args:
            file_content: 文件内容
            sheet_name: 工作表名称

        Returns:
            解析结果
        """
        try:
            wb = load_workbook(filename=file_content, read_only=True, data_only=True)
            ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

            return self._parse_sheet(ws, self.OUTPUT_COLUMN_MAPPING, 'outputKey')

        except InvalidFileException:
            return {
                'success': False,
                'outputs': [],
                'errors': [{'type': 'invalid_file', 'message': '无效的Excel文件格式'}],
                'warnings': []
            }
        except Exception as e:
            return {
                'success': False,
                'outputs': [],
                'errors': [{'type': 'parse_error', 'message': f'解析失败: {str(e)}'}],
                'warnings': []
            }

    def _parse_sheet(
        self,
        ws,
        column_mapping: Dict[str, str],
        key_field: str
    ) -> Dict[str, Any]:
        """
        解析工作表

        Args:
            ws: 工作表对象
            column_mapping: 列名映射
            key_field: 主键字段名

        Returns:
            解析结果
        """
        errors = []
        warnings = []
        items = []

        # 读取表头（第一行）
        headers = []
        header_mapping = {}  # 列索引 -> 标准字段名

        for col_idx, cell in enumerate(ws[1], start=1):
            raw_header = self._clean_cell_value(cell.value)
            if raw_header:
                headers.append(raw_header)
                # 查找映射
                normalized = raw_header.lower().strip()
                if normalized in column_mapping:
                    header_mapping[col_idx] = column_mapping[normalized]
                else:
                    # 未识别的列，保留原名
                    header_mapping[col_idx] = raw_header
                    warnings.append({
                        'type': 'unknown_column',
                        'column': raw_header,
                        'message': f'未识别的列名: {raw_header}'
                    })

        if not header_mapping:
            errors.append({'type': 'no_headers', 'message': '未找到有效的表头'})
            return {'success': False, 'items': [], 'errors': errors, 'warnings': warnings}

        # 检查是否有主键列
        has_key_field = key_field in header_mapping.values()
        if not has_key_field:
            errors.append({
                'type': 'missing_key_column',
                'message': f'缺少必需的列: {key_field}'
            })
            return {'success': False, 'items': [], 'errors': errors, 'warnings': warnings}

        # 读取数据行（从第2行开始）
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            item = {}
            has_data = False

            for col_idx, cell in enumerate(row, start=1):
                if col_idx in header_mapping:
                    field_name = header_mapping[col_idx]
                    value = self._clean_cell_value(cell.value)
                    if value is not None and value != '':
                        has_data = True
                    item[field_name] = value

            # 跳过空行
            if not has_data:
                continue

            # 验证主键
            key_value = item.get(key_field)
            if not key_value:
                warnings.append({
                    'type': 'missing_key',
                    'row': row_idx,
                    'message': f'第{row_idx}行缺少{key_field}'
                })
                continue

            items.append(item)

        return {
            'success': len(errors) == 0,
            'items': items,
            'errors': errors,
            'warnings': warnings,
            'rowCount': len(items)
        }

    def _clean_cell_value(self, value: Any) -> Any:
        """
        清理单元格值

        处理：
        1. 去除前后空格
        2. 处理编码问题
        3. 处理特殊字符

        Args:
            value: 原始值

        Returns:
            清理后的值
        """
        if value is None:
            return None

        # 字符串处理
        if isinstance(value, str):
            # 去除前后空格（包括全角空格）
            cleaned = value.strip().strip('\u3000')

            # 去除不可见字符
            cleaned = ''.join(char for char in cleaned if char.isprintable() or char in '\n\r\t')

            # 处理可能的编码问题
            try:
                # 尝试检测并修复编码
                if self._has_encoding_issue(cleaned):
                    cleaned = self._fix_encoding(cleaned)
            except Exception:
                pass  # 保持原值

            return cleaned if cleaned else None

        # 数值类型保持原样
        if isinstance(value, (int, float)):
            return value

        # 其他类型转字符串
        return str(value).strip()

    def _has_encoding_issue(self, text: str) -> bool:
        """检查是否有编码问题"""
        # 检查是否包含常见的乱码特征
        garbled_patterns = ['锟斤拷', '烫烫烫', '屯屯屯', '�']
        return any(p in text for p in garbled_patterns)

    def _fix_encoding(self, text: str) -> str:
        """尝试修复编码问题"""
        # 常见的编码修复尝试
        encodings = ['utf-8', 'gbk', 'gb2312', 'gb18030', 'big5']

        for src_enc in encodings:
            for dst_enc in encodings:
                if src_enc == dst_enc:
                    continue
                try:
                    fixed = text.encode(src_enc).decode(dst_enc)
                    if not self._has_encoding_issue(fixed):
                        return fixed
                except (UnicodeDecodeError, UnicodeEncodeError):
                    continue

        return text  # 无法修复，返回原值

    def detect_file_encoding(self, file_content: bytes) -> str:
        """
        检测文件编码

        Args:
            file_content: 文件字节内容

        Returns:
            编码名称
        """
        result = chardet.detect(file_content)
        encoding = result.get('encoding', 'utf-8')

        # 映射常见编码名称
        encoding_map = {
            'GB2312': 'gbk',
            'gb2312': 'gbk',
            'GBK': 'gbk',
            'GB18030': 'gb18030',
        }

        return encoding_map.get(encoding, encoding)


# 单例
excel_parser_service = ExcelParserService()
